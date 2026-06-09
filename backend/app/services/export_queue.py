from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models.entities import Invoice, Payment, Project


executor = ThreadPoolExecutor(max_workers=2)
tasks: dict[str, dict] = {}


def enqueue_export(project_ids: list[int] | None, export_types: list[str], fmt: str, db_factory, year: int | None = None, keyword: str | None = None, mode: str = "year", month: str | None = None) -> dict:
    task_id = f"export_{datetime.now():%Y%m%d}_{uuid4().hex[:8]}"
    tasks[task_id] = {"task_id": task_id, "status": "processing", "download_url": None}
    executor.submit(_run_export, task_id, project_ids, export_types, fmt, db_factory, year, keyword, mode, month)
    return tasks[task_id]


def get_export_status(task_id: str) -> dict | None:
    return tasks.get(task_id)


def get_export_path(task_id: str) -> Path | None:
    task = tasks.get(task_id)
    if not task or task.get("status") != "finished" or not task.get("path"):
        return None
    path = Path(task["path"])
    return path if path.exists() else None


def _date_bounds(year: int | None, mode: str = "year", month: str | None = None) -> tuple[datetime | None, datetime | None, date | None, date | None]:
    if mode == "month" and month:
        try:
            month_year, month_num = [int(part) for part in month.split("-", 1)]
            if 1 <= month_num <= 12:
                start_day = date(month_year, month_num, 1)
                end_day = date(month_year + 1, 1, 1) if month_num == 12 else date(month_year, month_num + 1, 1)
                return datetime.combine(start_day, datetime.min.time()), datetime.combine(end_day, datetime.min.time()), start_day, end_day
        except ValueError:
            pass
    if not year:
        return None, None, None, None
    return datetime(year, 1, 1), datetime(year + 1, 1, 1), date(year, 1, 1), date(year + 1, 1, 1)


def _keyword_condition(keyword: str):
    like = f"%{keyword.strip()}%"
    return or_(
        Project.project_no.like(like),
        Project.contract_no.like(like),
        Project.name.like(like),
        Project.customer.like(like),
        Project.party_a.like(like),
    )


def _run_export(task_id: str, project_ids: list[int] | None, export_types: list[str], fmt: str, db_factory, year: int | None = None, keyword: str | None = None, mode: str = "year", month: str | None = None) -> None:
    db: Session = db_factory()
    try:
        settings = get_settings()
        out_dir = settings.upload_dir / "exports"
        out_dir.mkdir(parents=True, exist_ok=True)
        path = out_dir / f"{task_id}.xlsx"
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        export_types = export_types or ["project"]
        start_dt, end_dt, start_day, end_day = _date_bounds(year, mode, month)
        keyword_condition = _keyword_condition(keyword) if keyword and keyword.strip() else None
        if "project" in export_types:
            ws = wb.create_sheet("项目明细")
            ws.append(["项目编号", "合同编号", "项目名称", "甲方/客户", "乙方", "合同金额", "状态", "尾款状态"])
            query = db.query(Project)
            if project_ids:
                query = query.filter(Project.id.in_(project_ids))
            if start_dt and end_dt:
                query = query.filter(Project.create_time >= start_dt, Project.create_time < end_dt)
            if keyword_condition is not None:
                query = query.filter(keyword_condition)
            for project in query.order_by(Project.create_time.desc()).all():
                ws.append([project.project_no, project.contract_no, project.name, project.party_a or project.customer, project.party_b, float(project.amount or 0), project.status, project.balance_status])
        if "invoice" in export_types:
            ws = wb.create_sheet("开票记录")
            ws.append(["项目编号", "项目名称", "发票号码", "不含税金额", "税率", "税额", "价税合计", "开票日期", "购方", "销方"])
            query = db.query(Invoice).join(Project)
            if project_ids:
                query = query.filter(Invoice.project_id.in_(project_ids))
            if start_day and end_day:
                query = query.filter(Invoice.invoice_date >= start_day, Invoice.invoice_date < end_day)
            if keyword_condition is not None:
                query = query.filter(keyword_condition)
            for invoice in query.order_by(Invoice.create_time.desc()).all():
                project = invoice.project
                ws.append([project.project_no if project else "", project.name if project else "", invoice.invoice_no, float(invoice.amount_without_tax or 0), float(invoice.tax_rate or 0), float(invoice.tax_amount or 0), float(invoice.amount or 0), invoice.invoice_date.isoformat(), invoice.buyer, invoice.seller])
        if "payment" in export_types:
            ws = wb.create_sheet("回款记录")
            ws.append(["项目编号", "项目名称", "关联发票", "回款金额", "回款日期", "方式", "凭证", "备注"])
            query = db.query(Payment).join(Project)
            if project_ids:
                query = query.filter(Payment.project_id.in_(project_ids))
            if start_day and end_day:
                query = query.filter(Payment.payment_date >= start_day, Payment.payment_date < end_day)
            if keyword_condition is not None:
                query = query.filter(keyword_condition)
            for payment in query.order_by(Payment.create_time.desc()).all():
                project = payment.project
                ws.append([project.project_no if project else "", project.name if project else "", payment.invoice.invoice_no if payment.invoice else "", float(payment.amount or 0), payment.payment_date.isoformat(), payment.payment_method, "已上传" if payment.voucher_file else "未上传", payment.remark])
        if not wb.sheetnames:
            ws = wb.create_sheet("空导出")
            ws.append(["提示"])
            ws.append(["未选择导出类型"])
        wb.save(path)
        tasks[task_id] = {"task_id": task_id, "status": "finished", "download_url": f"/api/v1/export/download?task_id={task_id}", "path": str(path), "file_name": f"{task_id}.xlsx"}
    except Exception as exc:
        tasks[task_id] = {"task_id": task_id, "status": "failed", "error": str(exc)}
    finally:
        db.close()
